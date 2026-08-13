#!/usr/bin/env python3
"""
Analyze ray_serve_test raw JSON output and produce a per-scenario summary including
latency metrics and pod resource metrics (CPU, memory, GPU).

Usage:
    ray_serve_test run --out json=results/raw.ndjson -e SUITE=rps_payload ... k6_ray_serve_test.js
    python analyze_k6_results.py results/raw.ndjson
    python analyze_k6_results.py results/raw.ndjson --out results/scenario_summary.json

Outputs:
    <out>.json   — full per-scenario summary
    <out>.xlsx   — Excel workbook with Latency and Resources sheets

Pod resource metrics (pod_cpu_millicores, pod_memory_mb, pod_gpu_util_pct,
pod_gpu_mem_used_mb) are shown per scenario when ENABLE_POD_METRICS=1 was set
during the ray_serve_test run. Each sample is tagged with the load scenario that was active
at poll time, so averages/p90/max reflect resource usage under that specific load.
"""

import json
import sys
import re
import argparse
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


RESOURCE_METRICS = [
    'pod_cpu_millicores',
    'pod_memory_mb',
    'pod_gpu_util_pct',
    'pod_gpu_mem_used_mb',
]

RESOURCE_LABELS = {
    'pod_cpu_millicores':  'CPU (cores)',
    'pod_memory_mb':       'Mem (GB)',
    'pod_gpu_util_pct':    'GPU (%)',
    'pod_gpu_mem_used_mb': 'GPU mem (GB)',
}


def percentile(sorted_values, p):
    if not sorted_values:
        return 0
    idx = (p / 100) * (len(sorted_values) - 1)
    lo, hi = int(idx), min(int(idx) + 1, len(sorted_values) - 1)
    return sorted_values[lo] + (sorted_values[hi] - sorted_values[lo]) * (idx - lo)


def parse_args():
    parser = argparse.ArgumentParser(description="Per-scenario ray_serve_test result analyzer")
    parser.add_argument("input", help="ray_serve_test raw NDJSON file (--out json=...)")
    parser.add_argument("--out", default=None, help="Output base path (no extension); defaults to input path stem")
    parser.add_argument("--metrics", nargs="+",
                        default=["mcp_c2_latency_ms", "prompt_c2_latency_ms"],
                        help="Latency metrics to analyze")
    return parser.parse_args()


def load_points(path, target_metrics):
    """Read NDJSON and collect data points grouped by (metric, scenario)."""
    data       = defaultdict(lambda: defaultdict(list))  # data[metric][scenario]
    errors     = defaultdict(int)                         # errors[scenario]
    timestamps = defaultdict(lambda: defaultdict(list))
    resources  = defaultdict(lambda: defaultdict(list))  # resources[metric][scenario]

    resource_set = set(RESOURCE_METRICS)

    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError:
                continue

            if obj.get("type") != "Point":
                continue

            metric     = obj.get("metric", "")
            point_data = obj.get("data", {})
            tags       = point_data.get("tags", {})
            scenario   = tags.get("scenario", "")
            value      = point_data.get("value", 0)
            ts         = point_data.get("time", "")

            if metric in target_metrics and scenario:
                data[metric][scenario].append(value)
                timestamps[metric][scenario].append(ts)

            elif metric in resource_set and scenario:
                if scenario not in ('pod_metrics',):
                    resources[metric][scenario].append(value)

            if metric == "error_rate" and scenario:
                if value == 1:
                    errors[scenario] += 1

    return data, errors, timestamps, resources


def _stat(vals):
    """Return a stats dict for a sorted list of values."""
    if not vals:
        return None
    sv = sorted(vals)
    return {
        "avg":     round(sum(sv) / len(sv), 1),
        "p50":     round(percentile(sv, 50), 1),
        "p90":     round(percentile(sv, 90), 1),
        "max":     round(max(sv), 1),
        "samples": len(sv),
    }


def build_summary(data, errors, timestamps, resources, target_metrics):
    all_scenarios = set()
    for metric in target_metrics:
        all_scenarios.update(data[metric].keys())
    for metric in RESOURCE_METRICS:
        all_scenarios.update(resources[metric].keys())

    results = {}
    for scenario in sorted(all_scenarios):
        entry = {"scenario": scenario, "metrics": {}, "resources": {}}

        m = re.search(r'_(\d+)b_rps(\d+)$', scenario)
        if m:
            size_bytes = int(m.group(1))
            target_rps = int(m.group(2))
            entry["payload_bytes"] = size_bytes
            entry["payload_label"] = f"{size_bytes // 1024}K" if size_bytes >= 1024 else f"{size_bytes}B"
            entry["target_rps"] = target_rps
        else:
            # Default/ad-hoc sweep: name ends with _<N>rps (e.g. prompt_c2_20rps)
            m2 = re.search(r'_(\d+)rps$', scenario)
            if m2:
                entry["target_rps"] = int(m2.group(1))

        entry["errors"] = errors.get(scenario, 0)

        for metric in target_metrics:
            values = sorted(data[metric].get(scenario, []))
            if not values:
                continue
            count = len(values)
            # Derive stage duration from the actual request timestamps so this
            # works regardless of whether stages are 15 s, 60 s, or 120 s.
            ts_list = timestamps[metric].get(scenario, [])
            if len(ts_list) >= 2:
                from datetime import datetime, timezone
                def _parse_ts(t):
                    return datetime.fromisoformat(t.replace('Z', '+00:00')).timestamp()
                try:
                    stage_duration_s = max(1, _parse_ts(max(ts_list)) - _parse_ts(min(ts_list)))
                except Exception:
                    stage_duration_s = 120
            else:
                stage_duration_s = 120
            achieved_rps = round(count / stage_duration_s, 2)

            entry["metrics"][metric] = {
                "count":        count,
                "achieved_rps": achieved_rps,
                "min":  round(min(values), 1),
                "p50":  round(percentile(values, 50), 1),
                "avg":  round(sum(values) / count, 1),
                "p90":  round(percentile(values, 90), 1),
                "p95":  round(percentile(values, 95), 1),
                "p99":  round(percentile(values, 99), 1),
                "max":  round(max(values), 1),
            }

        for rm in RESOURCE_METRICS:
            stat = _stat(resources[rm].get(scenario, []))
            if stat:
                entry["resources"][rm] = stat

        results[scenario] = entry

    return results


# ── Console output ─────────────────────────────────────────────────────────────

def print_combined_table(results, metric):
    rows = [(s, e) for s, e in results.items() if metric in e.get("metrics", {})]
    if not rows:
        return
    # Sort by payload size (bytes) ascending, then target RPS ascending.
    # Scenarios without a size/RPS in the name sort last.
    rows.sort(key=lambda x: (x[1].get("payload_bytes", float("inf")),
                              x[1].get("target_rps",   float("inf"))))

    print(f"\n── {metric} + pod resources per scenario {'─' * 40}")
    hdr = (f"{'Scenario':<38} {'payload':>7} {'target':>7} {'achiev':>7}"
           f" {'min':>7} {'avg':>7} {'p50':>7} {'p90':>7} {'p95':>7} {'p99':>7} {'max':>7} {'errors':>7}"
           f" {'CPUavg':>7} {'CPUp90':>7} {'CPUmax':>7}"
           f" {'Memavg':>7} {'Memp90':>7} {'Memmax':>7}"
           f" {'GPU%avg':>8} {'GPU%max':>8}"
           f" {'GMavg':>7} {'GMmax':>7}"
           f" {'smpl':>5}")
    print(hdr)
    print("─" * len(hdr))

    def fv(d, k, fmt=".0f", divisor=1):
        v = d.get(k)
        return format(v / divisor, fmt) if v is not None else "-"

    for scenario, entry in rows:
        m    = entry["metrics"][metric]
        res  = entry.get("resources", {})
        cpu  = res.get("pod_cpu_millicores", {})
        mem  = res.get("pod_memory_mb", {})
        gpu  = res.get("pod_gpu_util_pct", {})
        gmem = res.get("pod_gpu_mem_used_mb", {})
        samples = max((r.get("samples", 0) for r in res.values()), default=0) if res else "-"

        print(
            f"{scenario:<38}"
            f" {entry.get('payload_label', '-'):>7}"
            f" {entry.get('target_rps', '-'):>7}"
            f" {m['achieved_rps']:>7}"
            f" {m['min']:>7} {m['avg']:>7} {m['p50']:>7} {m['p90']:>7} {m['p95']:>7} {m['p99']:>7} {m['max']:>7}"
            f" {entry['errors']:>7}"
            f" {fv(cpu,  'avg'):>7} {fv(cpu,  'p90'):>7} {fv(cpu,  'max'):>7}"
            f" {fv(mem,  'avg', '.2f', 1024):>7} {fv(mem,  'p90', '.2f', 1024):>7} {fv(mem,  'max', '.2f', 1024):>7}"
            f" {fv(gpu,  'avg', '.1f'):>8} {fv(gpu,  'max', '.1f'):>8}"
            f" {fv(gmem, 'avg', '.2f', 1024):>7} {fv(gmem, 'max', '.2f', 1024):>7}"
            f" {samples!s:>5}"
        )
    print("─" * len(hdr))
    print("  CPU in millicores | Mem in GB | GPU% = utilization % | GMem in GB")


# ── Excel export ───────────────────────────────────────────────────────────────

def _xl_header_style():
    return {
        "font":      Font(bold=True, color="FFFFFF"),
        "fill":      PatternFill("solid", fgColor="1F4E79"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }

def _xl_subheader_style():
    return {
        "font":      Font(bold=True, color="FFFFFF"),
        "fill":      PatternFill("solid", fgColor="2E75B6"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }

def _apply(cell, styles):
    for attr, val in styles.items():
        setattr(cell, attr, val)

def _autofit(ws, min_width=10, max_width=50):
    for col in ws.columns:
        length = max(
            (len(str(cell.value)) if cell.value is not None else 0) for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_width), max_width)


def save_excel(results, target_metrics, xlsx_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    def endpoint_of(scenario):
        if scenario.startswith("prompt"):
            return "prompt-c2"
        if scenario.startswith("mcp"):
            return "mcp-c2"
        return "-"

    # Filter to metrics that actually have data in this result set
    active_metrics = [m for m in target_metrics
                      if any(m in entry.get("metrics", {}) for entry in results.values())]

    # ── Column layout ─────────────────────────────────────────────────────────
    # Fixed cols | Latency group per metric (only if data exists) | Resource groups
    fixed_cols      = ["Scenario", "Endpoint", "Payload", "Target RPS"]
    lat_sub_cols    = ["Count", "Achieved RPS", "Min (ms)", "p50 (ms)", "Avg (ms)",
                       "p90 (ms)", "p95 (ms)", "p99 (ms)", "Max (ms)", "Errors"]
    res_groups = [
        ("CPU (cores)",       "pod_cpu_millicores",  ["Avg", "p50", "p90", "Max"]),
        ("Memory (GB)",      "pod_memory_mb",        ["Avg", "p50", "p90", "Max"]),
        ("GPU Util (%)",     "pod_gpu_util_pct",     ["Avg", "p50", "p90", "Max"]),
        ("GPU Mem (GB)",     "pod_gpu_mem_used_mb",  ["Avg", "p50", "p90", "Max"]),
    ]

    # ── Row 1: group headers ──────────────────────────────────────────────────
    col = 1
    # Fixed columns span 2 rows
    for lbl in fixed_cols:
        cell = ws.cell(row=1, column=col, value=lbl)
        _apply(cell, _xl_header_style())
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1

    # Samples (single col, spans 2 rows, after fixed)
    samples_col = col
    cell = ws.cell(row=1, column=col, value="Samples")
    _apply(cell, _xl_header_style())
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
    col += 1

    # Latency group headers (only for metrics with data)
    for metric in active_metrics:
        label = "prompt-c2 latency (ms)" if "prompt" in metric else "mcp-c2 latency (ms)"
        end_col = col + len(lat_sub_cols) - 1
        cell = ws.cell(row=1, column=col, value=label)
        _apply(cell, _xl_subheader_style())
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=end_col)
        col = end_col + 1

    # Resource group headers
    for grp_label, _, sub_cols in res_groups:
        end_col = col + len(sub_cols) - 1
        cell = ws.cell(row=1, column=col, value=grp_label)
        _apply(cell, _xl_subheader_style())
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=end_col)
        col = end_col + 1

    # ── Row 2: sub-column headers ─────────────────────────────────────────────
    col = len(fixed_cols) + 2  # skip fixed (merged) + Samples (merged)
    for _ in active_metrics:
        for lbl in lat_sub_cols:
            cell = ws.cell(row=2, column=col, value=lbl)
            _apply(cell, _xl_header_style())
            col += 1
    for _, _, sub_cols in res_groups:
        for lbl in sub_cols:
            cell = ws.cell(row=2, column=col, value=lbl)
            _apply(cell, _xl_header_style())
            col += 1

    ws.freeze_panes = "A3"

    # Assign a distinct pastel fill to each unique payload size (sorted asc).
    # Scenarios without a payload size share a neutral color.
    _PAYLOAD_COLORS = [
        "DDEEFF",  # light blue
        "DDFFDD",  # light green
        "FFF3CC",  # light yellow
        "FFE0CC",  # light orange
        "F0DDFF",  # light purple
        "FFDDE6",  # light pink
        "CCFFFF",  # light cyan
        "EEFFCC",  # light lime
    ]
    _NO_PAYLOAD_COLOR = "F5F5F5"  # light grey for scenarios without a payload size

    unique_sizes = sorted({
        e.get("payload_bytes")
        for e in results.values()
        if e.get("payload_bytes") is not None
    })
    size_to_color = {
        sz: _PAYLOAD_COLORS[i % len(_PAYLOAD_COLORS)]
        for i, sz in enumerate(unique_sizes)
    }

    def _row_fill(entry):
        sz = entry.get("payload_bytes")
        hex_color = size_to_color.get(sz, _NO_PAYLOAD_COLOR)
        return PatternFill("solid", fgColor=hex_color)

    # ── Data rows (sorted by payload size asc, then RPS asc) ─────────────────
    for scenario, entry in sorted(results.items(),
                                  key=lambda x: (x[1].get("payload_bytes", float("inf")),
                                                 x[1].get("target_rps",   float("inf")))):
        res = entry.get("resources", {})
        samples = max((r.get("samples", 0) for r in res.values()), default=0) if res else None

        fixed_vals = [
            scenario,
            endpoint_of(scenario),
            entry.get("payload_label", "-"),
            entry.get("target_rps", "-"),
            samples,
        ]

        lat_vals = []
        for metric in active_metrics:
            m = entry["metrics"].get(metric, {})
            lat_vals += [
                m.get("count"),
                m.get("achieved_rps"),
                m.get("min"),
                m.get("p50"),
                m.get("avg"),
                m.get("p90"),
                m.get("p95"),
                m.get("p99"),
                m.get("max"),
                entry.get("errors", 0),
            ]

        res_vals = []
        for _, key, _ in res_groups:
            s = res.get(key, {})
            # Convert CPU from millicores to cores; memory metrics from MB to GB
            if key == "pod_cpu_millicores":
                divisor = 1000
            elif key in ("pod_memory_mb", "pod_gpu_mem_used_mb"):
                divisor = 1024
            else:
                divisor = 1
            res_vals += [
                round(s["avg"] / divisor, 3) if s.get("avg") is not None else None,
                round(s["p50"] / divisor, 3) if s.get("p50") is not None else None,
                round(s["p90"] / divisor, 3) if s.get("p90") is not None else None,
                round(s["max"] / divisor, 3) if s.get("max") is not None else None,
            ]

        ws.append(fixed_vals + lat_vals + res_vals)

        # Apply row background color based on payload size
        fill = _row_fill(entry)
        for cell in ws[ws.max_row]:
            cell.fill = fill

    _autofit(ws)
    wb.save(xlsx_path)
    return xlsx_path


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    args = parse_args()

    print(f"Reading {args.input} ...")
    data, errors, timestamps, resources = load_points(args.input, set(args.metrics))

    results = build_summary(data, errors, timestamps, resources, args.metrics)

    if not results:
        print("No per-scenario data found. Make sure you ran ray_serve_test with --out json=<file>")
        sys.exit(1)

    for metric in args.metrics:
        print_combined_table(results, metric)

    # Derive base output path (strip extension)
    base = args.out
    if base is None:
        base = (args.input
                .replace(".ndjson", "_scenario_summary")
                .replace(".json",   "_scenario_summary"))
    else:
        base = re.sub(r'\.(json|xlsx)$', '', base)

    json_path = base + ".json"
    with open(json_path, "w") as f:
        json.dump(results, f, indent=2)
    print(f"\nWrote {json_path}")

    if HAS_OPENPYXL:
        xlsx_path = base + ".xlsx"
        save_excel(results, args.metrics, xlsx_path)
        print(f"Wrote {xlsx_path}")
    else:
        print("Skipping Excel export — install openpyxl:  pip install openpyxl")


if __name__ == "__main__":
    main()
