#!/usr/bin/env python3
"""
Build Consolidated Multi-Tab Excel Workbook for PANW Jina Embeddings v2 Benchmarks on Cloud TPU v5e:
- Tab 1: prompt-c2 Onnx FP 32 v5e
- Tab 2: prompt-c2 Onnx FP 16 v5e
- Tab 3: Performance-per-Dollar Analysis
"""

import json, glob, os, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_tab(ws, summary_files, tab_title):
    ws.title = tab_title
    
    # Header styles
    font_main_hdr = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_main_hdr = PatternFill("solid", fgColor="1B365D")  # Deep Navy Blue
    
    font_sub_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    fill_sub_hdr = PatternFill("solid", fgColor="2C5282")   # Steel Blue
    
    font_data = Font(name="Calibri", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Define Layout
    fixed_cols = ["Scenario", "Endpoint", "Payload", "Target RPS", "Samples"]
    lat_sub_cols = ["Count", "Achieved RPS", "Min (ms)", "p50 (ms)", "Avg (ms)", "p90 (ms)", "p95 (ms)", "p99 (ms)", "Max (ms)", "Errors"]
    res_groups = [
        ("CPU (cores)", ["Avg", "p50", "p90", "Max"]),
        ("Memory (GB)", ["Avg", "p50", "p90", "Max"]),
        ("TPU Util (%)", ["Avg", "p50", "p90", "Max"]),
        ("TPU Mem (GB)", ["Avg", "p50", "p90", "Max"]),
    ]
    
    # Row 1: Main Group Headers
    col = 1
    for lbl in fixed_cols:
        cell = ws.cell(row=1, column=col, value=lbl)
        cell.font = font_main_hdr
        cell.fill = fill_main_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1
        
    # Latency Group Header
    lat_start = col
    lat_end = col + len(lat_sub_cols) - 1
    cell = ws.cell(row=1, column=lat_start, value="prompt-c2 latency (ms)")
    cell.font = font_main_hdr
    cell.fill = fill_main_hdr
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=lat_start, end_row=1, end_column=lat_end)
    col = lat_end + 1
    
    # Resource Group Headers
    for grp_lbl, sub_cols in res_groups:
        grp_start = col
        grp_end = col + len(sub_cols) - 1
        cell = ws.cell(row=1, column=grp_start, value=grp_lbl)
        cell.font = font_main_hdr
        cell.fill = fill_main_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=1, start_column=grp_start, end_row=1, end_column=grp_end)
        col = grp_end + 1

    # Row 2: Sub-headers
    col = len(fixed_cols) + 1
    for lbl in lat_sub_cols:
        cell = ws.cell(row=2, column=col, value=lbl)
        cell.font = font_sub_hdr
        cell.fill = fill_sub_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col += 1
        
    for _, sub_cols in res_groups:
        for lbl in sub_cols:
            cell = ws.cell(row=2, column=col, value=lbl)
            cell.font = font_sub_hdr
            cell.fill = fill_sub_hdr
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col += 1

    # Collect and Parse all records
    records = []
    for f in summary_files:
        with open(f, "r") as fp:
            data = json.load(fp)
        for sc, entry in data.items():
            if not isinstance(entry, dict) or "metrics" not in entry:
                continue
            
            # Determine payload and target RPS
            payload_bytes = entry.get("payload_bytes")
            payload_label = entry.get("payload_label")
            target_rps = entry.get("target_rps")
            
            # Fallback regex extraction if missing
            if payload_bytes is None or payload_label is None:
                if "1024b" in sc:
                    payload_bytes = 1024
                    payload_label = "1K"
                elif "2048b" in sc:
                    payload_bytes = 2048
                    payload_label = "2K"
                elif "5120b" in sc:
                    payload_bytes = 5120
                    payload_label = "5K"
                elif "7168b" in sc:
                    payload_bytes = 7168
                    payload_label = "7K"
                else:
                    payload_bytes = 0
                    payload_label = "-"
            
            if target_rps is None:
                m_rps = re.search(r'rps(\d+)', sc)
                if m_rps:
                    target_rps = int(m_rps.group(1))
                else:
                    # from filename
                    m_f = re.search(r'(\d+)rps', f)
                    target_rps = int(m_f.group(1)) if m_f else 0
            
            m = entry["metrics"].get("prompt_c2_latency_ms", {})
            res = entry.get("resources", {})
            samples = max((r.get("samples", 0) for r in res.values()), default=0) if res else 15
            
            records.append({
                "scenario": f"{sc}_rps{target_rps}" if f"rps{target_rps}" not in sc else sc,
                "endpoint": "prompt-c2",
                "payload_label": payload_label,
                "payload_bytes": payload_bytes,
                "target_rps": target_rps,
                "samples": samples,
                "count": m.get("count", 0),
                "achieved_rps": round(m.get("achieved_rps", 0), 2),
                "min": round(m.get("min", 0), 2),
                "p50": round(m.get("p50", 0), 2),
                "avg": round(m.get("avg", 0), 2),
                "p90": round(m.get("p90", 0), 2),
                "p95": round(m.get("p95", 0), 2),
                "p99": round(m.get("p99", 0), 2),
                "max": round(m.get("max", 0), 2),
                "errors": entry.get("errors", 0),
                "resources": res
            })
            
    # Sort records: Payload bytes asc, Target RPS asc
    records = sorted(records, key=lambda x: (x["payload_bytes"], x["target_rps"]))
    
    # Payload background colors
    PAYLOAD_COLORS = {
        1024: "EBF3FB",  # Soft Ice Blue
        2048: "EAF7EC",  # Soft Mint Green
        5120: "FEF9E7",  # Soft Warm Cream
        7168: "FDF2E9",  # Soft Peach
    }
    
    row_idx = 3
    for r in records:
        fill_color = PatternFill("solid", fgColor=PAYLOAD_COLORS.get(r["payload_bytes"], "FFFFFF"))
        
        row_vals = [
            r["scenario"],
            r["endpoint"],
            r["payload_label"],
            r["target_rps"],
            r["samples"],
            r["count"],
            r["achieved_rps"],
            r["min"],
            r["p50"],
            r["avg"],
            r["p90"],
            r["p95"],
            r["p99"],
            r["max"],
            r["errors"],
            # CPU
            "-", "-", "-", "-",
            # Mem
            "-", "-", "-", "-",
            # TPU Util
            "-", "-", "-", "-",
            # TPU Mem
            "-", "-", "-", "-"
        ]
        
        ws.append(row_vals)
        for col_idx in range(1, len(row_vals) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = font_data
            c.fill = fill_color
            c.border = thin_border
            if col_idx in [1, 2, 3]:
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.alignment = Alignment(horizontal="right", vertical="center")
                
            # Number formatting
            if col_idx in [7, 8, 9, 10, 11, 12, 13, 14]:
                c.number_format = '#,##0.00'
            elif col_idx in [4, 5, 6, 15]:
                c.number_format = '#,##0'
                
        row_idx += 1
        
    ws.freeze_panes = "A3"
    
    # Auto-fit column widths
    for col_cells in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col_cells[:2])
        data_max = max(len(str(cell.value or '')) for cell in col_cells[2:]) if len(col_cells) > 2 else 0
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, data_max + 3, 11)

def build_economics_tab(ws):
    ws.title = "Performance-$ Analysis"
    
    font_title = Font(name="Calibri", size=14, bold=True, color="1B365D")
    font_sec = Font(name="Calibri", size=12, bold=True, color="2C5282")
    font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    fill_hdr = PatternFill("solid", fgColor="1B365D")
    fill_subhdr = PatternFill("solid", fgColor="2C5282")
    font_data = Font(name="Calibri", size=10)
    font_bold = Font(name="Calibri", size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    fill_accent = PatternFill("solid", fgColor="EBF3FB")
    
    ws.append(["Palo Alto Networks: Jina Embeddings v2 Performance-per-Dollar & TCO Analysis"])
    ws.cell(row=1, column=1).font = font_title
    ws.append(["Strict Production SLA: P99 Round-Trip Latency < 50 ms across GKE Network"])
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=11, italic=True, color="555555")
    ws.append([])
    
    headers = [
        "Payload Size",
        "Architecture",
        "Precision",
        "Max Certified RPS (P99<50ms)",
        "P99 Latency at Max RPS",
        "1-Yr CUD Hourly Cost",
        "3-Yr CUD Hourly Cost",
        "RPS / $ (1-Yr CUD)",
        "RPS / $ (3-Yr CUD)",
        "Cost / 1M Embeddings (1-Yr)",
        "Cost / 1M Embeddings (3-Yr)",
        "Throughput Advantage",
        "Cost Savings (1-Yr CUD)"
    ]
    
    ws.append(headers)
    hdr_row = 4
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=hdr_row, column=c_idx)
        cell.font = font_hdr
        cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    rows = [
        # 1KB
        ["1 KB (1024B)", "NVIDIA L4 (g2-standard-4)", "FP32 / ONNX", 40, "22.80 ms", 0.45, 0.32, 88.9, 125.0, 3.125, 2.222, "1.0x (Baseline)", "0.0%"],
        ["1 KB (1024B)", "Cloud TPU v5e (ct5lp-1t)", "FP32", 190, "48.70 ms", 0.84, 0.54, 226.2, 351.9, 1.228, 0.789, "4.75x Throughput", "60.7% Savings"],
        ["1 KB (1024B)", "Cloud TPU v5e (ct5lp-1t)", "FP16 / BF16", 160, "28.90 ms", 0.84, 0.54, 190.5, 296.3, 1.458, 0.938, "4.00x Throughput", "53.3% Savings"],
        
        # 2KB
        ["2 KB (2048B)", "NVIDIA L4 (g2-standard-4)", "FP32 / ONNX", 40, "46.50 ms", 0.45, 0.32, 88.9, 125.0, 3.125, 2.222, "1.0x (Baseline)", "0.0%"],
        ["2 KB (2048B)", "Cloud TPU v5e (ct5lp-1t)", "FP32", 80, "26.30 ms", 0.84, 0.54, 95.2, 148.1, 2.917, 1.875, "2.00x Throughput", "6.7% Savings"],
        ["2 KB (2048B)", "Cloud TPU v5e (ct5lp-1t)", "FP16 / BF16", 80, "24.80 ms", 0.84, 0.54, 95.2, 148.1, 2.917, 1.875, "2.00x Throughput", "6.7% Savings"],
        
        # 5KB
        ["5 KB (5120B)", "NVIDIA L4 (g2-standard-4)", "FP32 / ONNX", 20, "49.10 ms", 0.45, 0.32, 44.4, 62.5, 6.250, 4.444, "1.0x (Baseline)", "0.0%"],
        ["5 KB (5120B)", "Cloud TPU v5e (ct5lp-1t)", "FP32", 80, "41.90 ms", 0.84, 0.54, 95.2, 148.1, 2.917, 1.875, "4.00x Throughput", "53.3% Savings"],
        ["5 KB (5120B)", "Cloud TPU v5e (ct5lp-1t)", "FP16 / BF16", 80, "39.50 ms", 0.84, 0.54, 95.2, 148.1, 2.917, 1.875, "4.00x Throughput", "53.3% Savings"],
        
        # 7KB
        ["7 KB (7168B)", "NVIDIA L4 (g2-standard-4)", "FP32 / ONNX", 10, "46.40 ms", 0.45, 0.32, 22.2, 31.3, 12.500, 8.889, "1.0x (Baseline)", "0.0%"],
        ["7 KB (7168B)", "Cloud TPU v5e (ct5lp-1t)", "FP32", 80, "49.10 ms", 0.84, 0.54, 95.2, 148.1, 2.917, 1.875, "8.00x Throughput", "76.7% Savings"],
        ["7 KB (7168B)", "Cloud TPU v5e (ct5lp-1t)", "FP16 / BF16", 80, "48.00 ms", 0.84, 0.54, 95.2, 148.1, 2.917, 1.875, "8.00x Throughput", "76.7% Savings"],
    ]
    
    for r_idx, row_data in enumerate(rows, start=5):
        ws.append(row_data)
        is_tpu = "TPU" in row_data[1]
        fill = fill_accent if is_tpu else PatternFill("solid", fgColor="FFFFFF")
        
        for c_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = font_bold if (is_tpu and c_idx in [4, 8, 9, 10, 11, 12, 13]) else font_data
            cell.fill = fill
            cell.border = thin_border
            if c_idx in [1, 2, 3, 12, 13]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                
            if c_idx in [6, 7, 10, 11]:
                cell.number_format = '$#,##0.000'
            elif c_idx in [8, 9]:
                cell.number_format = '#,##0.0'
            elif c_idx in [4]:
                cell.number_format = '#,##0'

    # Auto-fit column widths
    for col_cells in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def main():
    wb = openpyxl.Workbook()
    
    # Tab 1: FP32
    fp32_files = sorted(glob.glob("/workspace/cpu_to_tpu_results/*summary.json"))
    ws_fp32 = wb.active
    build_tab(ws_fp32, fp32_files, "prompt-c2 Onnx FP 32 v5e")
    
    # Tab 2: FP16
    fp16_files = sorted(glob.glob("/workspace/cpu_to_tpu_results_fp16/*summary.json"))
    ws_fp16 = wb.create_sheet()
    build_tab(ws_fp16, fp16_files, "prompt-c2 Onnx FP 16 v5e")
    
    # Tab 3: Performance-per-Dollar Analysis
    ws_econ = wb.create_sheet()
    build_economics_tab(ws_econ)
    
    out_path = "/workspace/jina_embeddings_v2_tpu_v5e_benchmarks.xlsx"
    wb.save(out_path)
    print(f"Successfully generated consolidated Excel workbook: {out_path}")

if __name__ == "__main__":
    main()
